import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def extract_values(query):
    values = {}
    
    # Split the query into parts by 'and'
    parts = re.split(r' and | or ', query)
    
    for part in parts:
        part = part.strip()

        if part.startswith('('):
            part = part[1:].strip()
        elif part.endswith(')'):
            part = part[:-1].strip()

        if ' IN ' in part and ' NOT IN ' not in part:
            # Handle 'IN' clauses
            key, value = part.split(' IN ')
            key = key.strip()
            value = value.strip().strip('()')
            values[key] = [v.strip().strip("'") for v in value.split(',')]
        elif ' NOT IN' in part:
            key, value = part.split(' NOT IN')
            key = key.strip()
            value = value.strip().strip('()')
            values[key] = [v.strip().strip("'") for v in value.split(',')]
            values['not_in_key'] = key
        else:
            # Handle key-value pairs
            key, value = part.split(' = ')
            key = key.strip()
            value = value.strip().strip("'")
            values[key] = value

    return values

# # Test extraction with a sample query
# sample_query = "subjectIdentifier.regulatoryRegimeIdentifier.name = 'The Monetary Authority Of Singapore 610' and transactionReportingStatus.transactionStateValue = 'Submitted' and reportableData.productId IN ('ForeignExchange:ComplexExotic','OTC Options:','ForeignExchange:SimpleExotic:Barrier','ForeignExchange:SimpleExotic:Digital','ForeignExchange:VanillaOption','ForeignExchange:NDO') and nonReportableData.bankType = 'COS' and nonReportableData.currencySection = 'SGD' and nonReportableData.businessDate = '2024-05-31T00:00:00.000Z' and nonReportableData.buyCurrency = 'USD'"
# extracted_values = extract_values(sample_query)
# print(extracted_values)

def filter_template(values, template_df, otc_options_col_index):
    
    filtered_data_section = pd.DataFrame()
    
    if 'subjectIdentifier.regulatoryRegimeIdentifier.name' in values:
        if template_df.iloc[7, 3] != 'Banks In Singapore':
            return None 

    if 'reportableData.productId' in values and 'OTC Options:' in values['reportableData.productId']:
        if 'nonReportableData.currencySection' in values:
            if values['nonReportableData.currencySection'] == 'SGD':
                if template_df.iloc[9, 0] != 'SGD against':
                    return None  
                filtered_data_section = template_df.iloc[10:30, :]
            elif values['nonReportableData.currencySection'] == 'USD':
                if template_df.iloc[30, 0] != 'USD against':
                    return None  
                filtered_data_section = template_df.iloc[31:56, :]
            elif values['nonReportableData.currencySection'] == 'EUR':
                if template_df.iloc[58, 0] != 'EUR against':
                    return None  
                filtered_data_section = template_df.iloc[59:64, :]
            elif values['nonReportableData.currencySection'] == 'JPY':
                if template_df.iloc[66, 0] != 'JPY against':
                    return None 
                filtered_data_section = template_df.iloc[67:73, :]
            elif values['nonReportableData.currencySection'] == 'Other':
                if template_df.iloc[75, 0] != 'All other currency pairs against':
                    return None 
                filtered_data_section = template_df.iloc[76:79, :]
            else:
                return 0 

    if 'nonReportableData.buyCurrency' in values:
        buy_currency = values['nonReportableData.buyCurrency']
    elif 'nonReportableData.sellCurrency' in values:
        buy_currency = values['nonReportableData.sellCurrency']
    else:
        buy_currency = None
    
    not_in_key = values.get('not_in_key')
    not_in_currency = values.get(not_in_key, []) if not_in_key else []

    total_value = 0

    if not_in_currency:
        all_currencies = set(filtered_data_section.iloc[:, 1].dropna().unique())
        if '  of which:' in all_currencies:
            all_currencies.remove('  of which:')
            all_currencies.update(['CNH', 'CNY'])
        excluded_currencies = set(not_in_currency)
        included_currencies = all_currencies - excluded_currencies
        for currency in included_currencies:
            for idx, row in filtered_data_section.iterrows():
                if row[1] == '  of which:' or pd.isna(row[1]):
                    if row[2] == currency:
                        total_value += row[otc_options_col_index]
                elif row[1] == currency:
                    total_value += row[otc_options_col_index]
        return total_value

    if buy_currency:
        if isinstance(buy_currency, list):
            for buy_currency in buy_currency:
                for idx, row in filtered_data_section.iterrows():
                    if row[1] == '  of which:' or pd.isna(row[1]):
                        if row[2] == buy_currency:
                            total_value += row[otc_options_col_index]
                    elif row[1] == buy_currency:
                        total_value += row[otc_options_col_index]
            return total_value
            
        else:
            for idx, cell in filtered_data_section.iterrows():
                if cell[1] == '  of which:' or pd.isna(cell[1]):
                    if cell[2] == buy_currency:
                        return cell[otc_options_col_index]
                elif cell[1] == buy_currency:
                    return cell[otc_options_col_index]
                
    return total_value if total_value > 0 else 0

# Load the Aggregate and Tenplate Excel files
aggregate_df = pd.read_excel(r'C:\\Links\\Joydeeep Code\\PythonCode\\AggregationFiles\\Aggregation_queries.xlsx', sheet_name='Sheet1')
template_df = pd.read_excel(r'C:\\Links\\Joydeeep Code\\PythonCode\\AggregationFiles\\Tenplate.xlsx', sheet_name='I_Pt I_FX Turnover (M)')

otc_columns = [
    ("Banks In Singapore", 7),
    ("Banks Outside Singapore", 28),
    ("Non-Financial Customers Outside Singapore", 42)
]

for name, column in otc_columns:
    results_col = f"Result for {name}"
    values_col = f"OTC value for {name}"
    results = []
    values_in_template = []

    for index, row in aggregate_df.iterrows():
        query = row['Test Case Query']
        expected_result = row['Report Result']

        extracted_values = extract_values(query)

        filtered_data = filter_template(extracted_values, template_df, otc_options_col_index=column)
        values_in_template.append(filtered_data)

        if filtered_data == expected_result:
            results.append('Pass')
        else:
            results.append('Fail')

    aggregate_df[values_col] = values_in_template
    aggregate_df[results_col] = results

output_path = r'C:\\Links\\Joydeeep Code\\PythonCode\\AggregationFiles\\Updated_Aggregate.xlsx'
aggregate_df.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active

result_columns = [f"G", f"I", f"K"]
for col in result_columns:
    for idx, cell in enumerate(ws[col][1:], 1):
        if cell.value == 'Pass':
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Set column widths and wrap text for 'Test Case Query' column
column_widths = {
    'A': 10,
    'B': 70,
    'C': 55,
    'D': 20,
    'E': 20,
    'F': 30,
    'G': 30,
    'H': 35,
    'I': 35,
    'J': 50,
    'K': 50,
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Enable text wrapping for 'Test Case Query' column
for cell in ws['B']:
    cell.alignment = Alignment(wrap_text=True)

wb.save(output_path)
